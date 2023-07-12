import openpyxl


class HomeLoginData:

    def get_data_xlsx(self):
        book = openpyxl.load_workbook(r"D:\Documents\Trip - Copy\excel_data\sheet1.xlsx")
        sheet = book.active
        use = sheet.cell(row=1, column=2).value
        pas = sheet.cell(row=1, column=3).value

        title = [use, pas]

        lis = []
        for r in range(2, sheet.max_row+1):
            dit = {}
            i = 0
            for c in range(2, sheet.max_column+1):
                print(sheet.cell(row=r, column=c).value)
                dit[title[i]] = sheet.cell(row=r, column=c).value
                i += 1
            lis.append(dit)
        return lis












